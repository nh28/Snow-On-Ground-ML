import os
import pandas as pd
import openpyxl

excel_file_path = os.path.abspath("SOG.xlsx") 
excel_sheet = "Sheet1"
df = pd.read_excel(excel_file_path, sheet_name=excel_sheet)

workbook = openpyxl.load_workbook('SOG Summer GT 0 pivot Phase 3.xlsx')
station_list = workbook['Phase3Stations']
stations = []
for station_row in station_list.iter_rows(min_row=2, max_row=station_list.max_row, min_col=2, max_col=2):
        for station_cell in station_row:
            stations.append(station_cell.value)

for index, station_row in df.iterrows():
    station_id = station_row['STN_ID']
    type = station_row['MEAS_TYPE_ID']
    year = station_row['LOCAL_YEAR']
    month = station_row['LOCAL_MONTH']
    day = station_row['LOCAL_DAY']

    if month in [5,6]:
        pivot_worksheet = workbook['Pivot 5,6']
    elif month in [7, 8]:
        pivot_worksheet = workbook['Pivot 7,8']
    else:
        pivot_worksheet = workbook['Pivot 9']
    
    
    for station_idx, station_row in enumerate(pivot_worksheet.iter_rows(min_row=3, max_row=pivot_worksheet.max_row, min_col=1, max_col=1)):
        for station_cell in station_row:
            if station_cell.value == station_id:
                start_station_idx = station_idx + 4
                end_station_idx = station_idx + 4
                found_station = False
                while found_station == False and end_station_idx <= pivot_worksheet.max_row:
                    value_station = pivot_worksheet.cell(row= end_station_idx, column = 1).value
                    if value_station in stations:
                        found_station = True
                        end_station_idx-=1
                    else:
                        end_station_idx += 1

                for program_idx, program_row in enumerate(pivot_worksheet.iter_rows(min_row=start_station_idx, max_row=end_station_idx, min_col=1, max_col=1)):
                    for program_cell in program_row:
                        if program_cell.value == type:
                            start_program_idx = program_idx + start_station_idx
                            end_program_idx = program_idx + start_station_idx + 1

                            found_program = False
                            while found_program == False and end_program_idx <= end_station_idx - 1:
                                value_program = pivot_worksheet.cell(row= end_program_idx, column = 1).value
                                if value_program in [951, 15498, 14910]:
                                    found_program = True
                                    end_program_idx -= 1
                                else:
                                    end_program_idx += 1
                            
                            for year_idx, year_row in enumerate(pivot_worksheet.iter_rows(min_row=start_program_idx, max_row=end_program_idx, min_col=1, max_col=1)):
                                for year_cell in year_row:
                                    if year_cell.value == year:
                                        start_year_idx = year_idx + start_program_idx
                                        end_year_idx = year_idx + start_program_idx + 1

                                        found_year = False
                                        while found_year == False and end_year_idx <= end_program_idx - 1:
                                            value_year = pivot_worksheet.cell(row= end_year_idx, column = 1).value
                                            if value_year in range(1991,2024):
                                                found_year = True
                                                end_year_idx -= 1
                                            else:
                                                end_year_idx += 1
                                        
                                        for month_idx, month_row in enumerate(pivot_worksheet.iter_rows(min_row=start_year_idx, max_row=end_year_idx, min_col=1, max_col=1)):
                                            for month_cell in month_row:
                                                if month_cell.value == month:
                                                    start_month_idx = month_idx + start_year_idx
                                                    col = day + 1
                                                    
                                                    cell = pivot_worksheet.cell(row= start_month_idx, column = col)
                                                    if cell.fill.start_color.index == 'FF92D050':
                                                        df.at[index, 'Accurate'] = 1
                                                    else:
                                                         df.at[index, 'Accurate'] = 0

df.to_excel(excel_file_path, sheet_name=excel_sheet, index=False)

                                    
                                        



                
